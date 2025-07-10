import json
from typing import Dict, List, Tuple

from openpyxl import load_workbook

PERIODS = ["E", "S", "A", "B", "C"]
MAX_STUDENTS_PER_TEACHER = 2


def parse_schedule(text: str) -> Dict[str, List[str]]:
    """Parse schedule string like '7/21(ABC),7/22(ES)'."""
    result: Dict[str, List[str]] = {}
    if not text:
        return result
    for item in str(text).split(','):
        item = item.strip()
        if not item or '(' not in item:
            continue
        day_part, period_part = item.split('(', 1)
        periods = [p for p in period_part.rstrip(')') if p in PERIODS]
        if not periods:
            continue
        result.setdefault(day_part, []).extend(periods)
    return result


def parse_subjects(text: str) -> Dict[str, int]:
    """Parse subject counts like 'Math(2),Science(1)'."""
    result: Dict[str, int] = {}
    if not text:
        return result
    for item in str(text).split(','):
        item = item.strip()
        if not item:
            continue
        if '(' in item and item.endswith(')'):
            name, count = item[:-1].split('(')
            result[name] = int(count)
        else:
            result[item] = 1
    return result


class Teacher:
    def __init__(self, name: str, schedule: Dict[str, List[str]], subjects: List[str]):
        self.name = name
        self.schedule = schedule
        self.subjects = [s.strip() for s in subjects if s.strip()]


class Student:
    def __init__(self, name: str, grade: str, availability: Dict[str, List[str]],
                 desired_count: int, desired_subjects: Dict[str, int]):
        self.name = name
        self.grade = grade
        self.availability = availability
        self.desired_count = int(desired_count)
        self.desired_subjects = desired_subjects


def load_teachers(filename: str) -> List[Teacher]:
    wb = load_workbook(filename)
    ws = wb.active
    teachers: List[Teacher] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        name = str(row[0]).strip()
        schedule = parse_schedule(row[1])
        subjects = str(row[2]).split(',') if row[2] else []
        teachers.append(Teacher(name, schedule, subjects))
    return teachers


def load_students(filename: str) -> List[Student]:
    wb = load_workbook(filename)
    ws = wb.active
    students: List[Student] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        name = str(row[0]).strip()
        grade = str(row[1]).strip() if row[1] else ''
        availability = parse_schedule(row[2])
        desired_count = int(row[3]) if row[3] else 0
        desired_subjects = parse_subjects(row[4])
        students.append(Student(name, grade, availability, desired_count, desired_subjects))
    return students


def create_schedule(teachers: List[Teacher], students: List[Student]) -> Dict[str, Dict[str, Dict[str, List[Tuple[str, str]]]]]:
    """Return schedule[day][period][teacher] = [(student, subject), ...]."""
    schedule: Dict[str, Dict[str, Dict[str, List[Tuple[str, str]]]]] = {}
    # place teachers
    for t in teachers:
        for day, periods in t.schedule.items():
            day_table = schedule.setdefault(day, {p: {} for p in PERIODS})
            for p in periods:
                day_table[p].setdefault(t.name, [])

    # assign students
    for s in students:
        # flatten available slots
        slots: List[Tuple[str, str]] = []
        for day, periods in s.availability.items():
            for p in periods:
                slots.append((day, p))
        used_slots: set[Tuple[str, str]] = set()
        for subject, need in s.desired_subjects.items():
            assigned = 0
            for day, period in slots:
                if assigned >= need or (day, period) in used_slots:
                    continue
                # find teacher
                for t in teachers:
                    if subject not in t.subjects:
                        continue
                    if day not in t.schedule or period not in t.schedule[day]:
                        continue
                    current = schedule[day][period].setdefault(t.name, [])
                    if len(current) >= MAX_STUDENTS_PER_TEACHER:
                        continue
                    current.append((s.name, subject))
                    used_slots.add((day, period))
                    assigned += 1
                    break
    return schedule


def print_schedule(schedule: Dict[str, Dict[str, Dict[str, List[Tuple[str, str]]]]]):
    for day in sorted(schedule.keys()):
        print(day)
        for period in PERIODS:
            print(f"  [{period}]")
            teachers = schedule[day].get(period, {})
            if not teachers:
                print("    (no teachers)")
                continue
            for teacher, students in teachers.items():
                if students:
                    stu_str = ', '.join(f"{n}({sub})" for n, sub in students)
                else:
                    stu_str = '---'
                print(f"    {teacher}: {stu_str}")
        print()


def save_schedule(schedule: Dict[str, Dict[str, Dict[str, List[Tuple[str, str]]]]], filename: str = 'schedule.json'):
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(schedule, f, ensure_ascii=False, indent=2)
    print(f"Schedule saved to {filename}")


def main():
    teachers = load_teachers('teachers.xlsx')
    students = load_students('students.xlsx')
    schedule = create_schedule(teachers, students)
    print_schedule(schedule)
    save_schedule(schedule)


if __name__ == '__main__':
    main()
